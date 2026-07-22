import os
import json
import math
import csv
import shutil
import openpyxl
from openpyxl.styles import PatternFill
import difflib
import re

# ==========================================
# CONFIGURATION
# ==========================================
HYBRID_FILE = "cpe_data_hybrid.xlsx"
OUTPUT_FILE = "cpe_data_hybrid_analyzed.xlsx"
ISSUES_FILE = "cpe_data_hybrid_issues.xlsx"
TOWERS_JSON_FILE = "towers_cache.json"
SUBNET_CSV_FILE = "subnets.csv"
RF_MASTER_FILE = "sector_details_complete.csv"

# ==========================================
# TRIGONOMETRY LOGIC
# ==========================================
def calculate_initial_compass_bearing(pointA, pointB):
    lat1, lon1 = math.radians(pointA[0]), math.radians(pointA[1])
    lat2, lon2 = math.radians(pointB[0]), math.radians(pointB[1])
    diffLong = lon2 - lon1
    x = math.sin(diffLong) * math.cos(lat2)
    y = math.cos(lat1) * math.sin(lat2) - (math.sin(lat1) * math.cos(lat2) * math.cos(diffLong))
    initial_bearing = math.atan2(x, y)
    initial_bearing = math.degrees(initial_bearing)
    return (initial_bearing + 360) % 360

def is_cpe_in_sector(cpe_bearing, sector_azimuth, sector_beamwidth):
    diff = abs(cpe_bearing - sector_azimuth)
    angular_difference = min(diff, 360 - diff)
    return angular_difference <= (sector_beamwidth / 2)

def get_angular_difference(bearing, azimuth):
    diff = abs(bearing - azimuth)
    return min(diff, 360 - diff)

# ==========================================
# DIRTY DATA HELPERS (Normalization & Fuzzy)
# ==========================================
def clean_str(val):
    if val is None: return ""
    return str(val).strip().upper()

def is_empty(val):
    return clean_str(val) in ["", "N/A", "NONE", "NAN", "NULL"]

def fuzzy_match_sector(dirty_name, available_sectors, threshold=0.85):
    if not dirty_name or is_empty(dirty_name): return None
    dirty_name = clean_str(dirty_name)
    
    # 1. Try Exact Match
    if dirty_name in available_sectors:
        return dirty_name
        
    # 2. Try Fuzzy Match
    matches = difflib.get_close_matches(dirty_name, available_sectors, n=1, cutoff=threshold)
    if matches:
        return matches[0]
        
    return None

def parse_float(val):
    if is_empty(val): return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

# ==========================================
# DATA LOADERS
# ==========================================
def load_towers():
    if not os.path.exists(TOWERS_JSON_FILE): return {}
    with open(TOWERS_JSON_FILE, "r") as f:
        data = json.load(f)
        # Normalize tower names
        return {clean_str(item["name"]): item for item in data}

def load_subnets():
    subnet_map = {}
    if os.path.exists(SUBNET_CSV_FILE):
        with open(SUBNET_CSV_FILE, "r") as f:
            reader = csv.reader(f)
            # Skip header
            next(reader, None)
            for row in reader:
                if len(row) >= 2:
                    site = row[0].strip()
                    ip_pattern = row[1].strip()
                    # Extract the 3rd octet (e.g. 10 from 172.16.10.xx)
                    match = re.search(r'172\.16\.(\d+)\.', ip_pattern)
                    if match and site:
                        base_octet = match.group(1)
                        subnet_map[base_octet] = clean_str(site)
    return subnet_map

def get_site_from_ip(ip, subnet_map):
    if is_empty(ip): return None
    parts = str(ip).split('.')
    if len(parts) == 4:
        return subnet_map.get(parts[2])
    return None

def load_rf_master():
    sector_data = {}
    site_sectors = {}
    
    if not os.path.exists(RF_MASTER_FILE):
        print(f"[-] ERROR: Master RF file not found at {RF_MASTER_FILE}")
        return sector_data, site_sectors
        
    try:
        temp_file = "temp_rf_master.xlsx"
        shutil.copy2(RF_MASTER_FILE, temp_file)
        wb = openpyxl.load_workbook(temp_file, data_only=True)
        ws = wb.active
        
        headers = [clean_str(cell.value).lower() for cell in ws[1]]
        
        idx_site = headers.index('site name') if 'site name' in headers else 0
        idx_az = headers.index('azimut') if 'azimut' in headers else 3
        idx_bw = headers.index('beam width') if 'beam width' in headers else 4
        idx_name = headers.index('sector name') if 'sector name' in headers else 6
        
        for row in ws.iter_rows(min_row=2):
            site = clean_str(row[idx_site].value)
            azimuth = parse_float(row[idx_az].value)
            beamwidth = parse_float(row[idx_bw].value)
            sector_name = clean_str(row[idx_name].value)
            
            if not is_empty(sector_name) and azimuth is not None:
                bw = beamwidth if beamwidth is not None else 90.0
                sector_data[sector_name] = {
                    "azimuth": azimuth,
                    "beamwidth": bw,
                    "site": site
                }
                if not is_empty(site):
                    if site not in site_sectors:
                        site_sectors[site] = []
                    site_sectors[site].append({
                        "name": sector_name,
                        "azimuth": azimuth,
                        "beamwidth": bw
                    })
                    
        wb.close()
        if os.path.exists(temp_file): os.remove(temp_file)
    except Exception as e:
        print(f"[-] Failed to load RF Master file: {e}")
        
    return sector_data, site_sectors

# ==========================================
# MAIN LOGIC
# ==========================================
if __name__ == "__main__":
    print("[*] Loading Master Databases...")
    towers_map = load_towers()
    subnet_map = load_subnets()
    sector_data, site_sectors = load_rf_master()
    
    valid_sectors_list = list(sector_data.keys())
    
    print(f"[+] Loaded {len(towers_map)} towers.")
    print(f"[+] Loaded {len(sector_data)} sector RF configurations.")
    
    if not os.path.exists(HYBRID_FILE):
        print(f"[-] ERROR: Hybrid file {HYBRID_FILE} not found!")
        exit(1)
        
    print(f"\n[*] Analyzing {HYBRID_FILE} with Dirty Data checks...")
    wb = openpyxl.load_workbook(HYBRID_FILE)
    ws = wb.active
    
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    
    new_cols = [
        'CPE Bearing', 'cnMaestro Sector Azimuth', 'cnMaestro Sector Beamwidth', 
        'Facing cnMaestro Sector (Yes/No)', 'Facing IPAM Sector (Yes/No)', 
        'Ideal Sector', 'IPAM Sector in DB (Yes/No)', 'cnMaestro AP in DB (Yes/No)',
        'Fuzzy Match Used (cnMaestro/IPAM)'
    ]
    for col in new_cols:
        ws.cell(row=1, column=len(headers) + 1).value = col
        headers.append(col)
        
    idx_ip = headers.index('IP Address') + 1
    idx_ap = headers.index('cnMaestro AP Name') + 1
    idx_ipam_sec = headers.index('IPAM Sector') + 1 if 'IPAM Sector' in headers else -1
    idx_lat = headers.index('IPAM SM Latitude') + 1
    idx_lon = headers.index('IPAM SM Longitude') + 1
    
    idx_dist_diff = headers.index('Distance Difference (km)') + 1 if 'Distance Difference (km)' in headers else -1
    
    wb_issues = openpyxl.Workbook()
    ws_issues = wb_issues.active
    ws_issues.append(headers)
    
    col_bearing = len(headers) - 8
    col_azimuth = len(headers) - 7
    col_beamwidth = len(headers) - 6
    col_facing_cnm = len(headers) - 5
    col_facing_ipam = len(headers) - 4
    col_ideal = len(headers) - 3
    col_ipam_db = len(headers) - 2
    col_cnm_db = len(headers) - 1
    col_fuzzy = len(headers)
    
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    processed = 0
    fuzzy_fixes = 0
    not_facing = 0
    
    for row in range(2, ws.max_row + 1):
        ip = clean_str(ws.cell(row=row, column=idx_ip).value)
        if is_empty(ip): continue
        
        raw_ap_name = clean_str(ws.cell(row=row, column=idx_ap).value)
        raw_ipam_sec = clean_str(ws.cell(row=row, column=idx_ipam_sec).value) if idx_ipam_sec != -1 else ""
        
        cpe_lat = parse_float(ws.cell(row=row, column=idx_lat).value)
        cpe_lon = parse_float(ws.cell(row=row, column=idx_lon).value)
        
        cpe_bearing = "N/A"
        sec_az = "N/A"
        sec_bw = "N/A"
        is_facing_cnm = "N/A"
        is_facing_ipam = "N/A"
        ideal_sector = "N/A"
        fuzzy_note = ""
        
        # Fuzzy Match CNM Sector
        matched_cnm = fuzzy_match_sector(raw_ap_name, valid_sectors_list)
        if matched_cnm and matched_cnm != raw_ap_name:
            fuzzy_note += f"CNM:{matched_cnm} "
            fuzzy_fixes += 1
            
        # Fuzzy Match IPAM Sector
        matched_ipam = fuzzy_match_sector(raw_ipam_sec, valid_sectors_list)
        if matched_ipam and matched_ipam != raw_ipam_sec:
            fuzzy_note += f"IPAM:{matched_ipam} "
            fuzzy_fixes += 1
            
        if cpe_lat is not None and cpe_lon is not None:
            try:
                rf_info_cnm = sector_data.get(matched_cnm) if matched_cnm else None
                site = rf_info_cnm["site"] if rf_info_cnm else get_site_from_ip(ip, subnet_map)
                
                if site and site in towers_map:
                    t_lat = parse_float(towers_map[site]["lat"])
                    t_lon = parse_float(towers_map[site]["lon"])
                    
                    if t_lat is not None and t_lon is not None:
                        bearing = calculate_initial_compass_bearing((t_lat, t_lon), (cpe_lat, cpe_lon))
                        cpe_bearing = round(bearing, 2)
                        
                        if rf_info_cnm:
                            sec_az = rf_info_cnm["azimuth"]
                            sec_bw = rf_info_cnm["beamwidth"]
                            if is_cpe_in_sector(bearing, sec_az, sec_bw):
                                is_facing_cnm = "Yes"
                            else:
                                is_facing_cnm = "No"
                                not_facing += 1
                                
                        rf_info_ipam = sector_data.get(matched_ipam) if matched_ipam else None
                        if rf_info_ipam:
                            if is_cpe_in_sector(bearing, rf_info_ipam["azimuth"], rf_info_ipam["beamwidth"]):
                                is_facing_ipam = "Yes"
                            else:
                                is_facing_ipam = "No"
                                
                        if site in site_sectors and len(site_sectors[site]) > 0:
                            closest_sector = None
                            min_diff = 360.0
                            for s in site_sectors[site]:
                                d = get_angular_difference(bearing, s["azimuth"])
                                if d < min_diff:
                                    min_diff = d
                                    closest_sector = s["name"]
                            ideal_sector = closest_sector
            except Exception:
                pass
                
        # Write Base Data
        ws.cell(row=row, column=col_bearing).value = cpe_bearing
        ws.cell(row=row, column=col_azimuth).value = sec_az
        ws.cell(row=row, column=col_beamwidth).value = sec_bw
        ws.cell(row=row, column=col_ideal).value = ideal_sector
        
        # Formatting Facing Statuses
        f_cnm = ws.cell(row=row, column=col_facing_cnm)
        f_cnm.value = is_facing_cnm
        if is_facing_cnm == "Yes": f_cnm.fill = green_fill
        elif is_facing_cnm == "No": f_cnm.fill = red_fill
            
        f_ipam = ws.cell(row=row, column=col_facing_ipam)
        f_ipam.value = is_facing_ipam
        if is_facing_ipam == "Yes": f_ipam.fill = green_fill
        elif is_facing_ipam == "No": f_ipam.fill = red_fill
            
        # Formatting Database Checks
        in_db_ipam = "Yes" if matched_ipam else "No"
        in_db_cnm = "Yes" if matched_cnm else "No"
        
        cell_ipam = ws.cell(row=row, column=col_ipam_db)
        cell_ipam.value = in_db_ipam
        if in_db_ipam == "Yes": cell_ipam.fill = green_fill
        else: cell_ipam.fill = red_fill
            
        cell_cnm = ws.cell(row=row, column=col_cnm_db)
        cell_cnm.value = in_db_cnm
        if in_db_cnm == "Yes": cell_cnm.fill = green_fill
        else: cell_cnm.fill = red_fill
        
        # Fuzzy Match Note
        cell_fuzzy = ws.cell(row=row, column=col_fuzzy)
        cell_fuzzy.value = fuzzy_note.strip() if fuzzy_note else "Exact Match"
        if fuzzy_note:
            cell_fuzzy.fill = yellow_fill
            
        # Write to Issues Sheet
        has_issue = False
        if is_facing_cnm == "No" or is_facing_ipam == "No": has_issue = True
        
        if idx_dist_diff != -1:
            d_val = parse_float(ws.cell(row=row, column=idx_dist_diff).value)
            if d_val and d_val > 1.0: has_issue = True
                
        if has_issue:
            ws_issues.append([ws.cell(row=row, column=c).value for c in range(1, len(headers) + 1)])
            r_idx = ws_issues.max_row
            if is_facing_cnm == "No": ws_issues.cell(row=r_idx, column=col_facing_cnm).fill = red_fill
            if is_facing_ipam == "No": ws_issues.cell(row=r_idx, column=col_facing_ipam).fill = red_fill
            if in_db_ipam == "No": ws_issues.cell(row=r_idx, column=col_ipam_db).fill = red_fill
            if in_db_cnm == "No": ws_issues.cell(row=r_idx, column=col_cnm_db).fill = red_fill
            if idx_dist_diff != -1 and d_val and d_val > 1.0:
                ws_issues.cell(row=r_idx, column=idx_dist_diff).fill = red_fill

        processed += 1

    try:
        wb.save(OUTPUT_FILE)
        wb_issues.save(ISSUES_FILE)
        print(f"\n[+] Analysis Complete! Processed {processed} CPEs.")
        print(f"    -> Fuzzy Rescued Sectors: {fuzzy_fixes}")
        print(f"    -> Connected to WRONG Sector: {not_facing}")
        print(f"    -> Identified Issues saved to {ISSUES_FILE} (Rows: {ws_issues.max_row - 1})")
        print(f"[+] Saved full results to {OUTPUT_FILE}")
    except PermissionError:
        print(f"\n[-] ERROR: Permission denied! Please close {OUTPUT_FILE} if it is open in Excel and run the script again.")
